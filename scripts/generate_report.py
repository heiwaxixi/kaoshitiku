from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]
REPORT_PATH = BASE_DIR / "output" / "reports" / "考试AI题库_交付说明_20260702.xlsx"

GREEN = "0F8F69"
GREEN_LIGHT = "DDF7EC"
BLUE = "2563EB"
BLUE_LIGHT = "EAF2FF"
AMBER_LIGHT = "FFF4DE"
RED_LIGHT = "FFE8E6"
GRAY_LIGHT = "F4F7F6"
BORDER = "D9E4E0"


def style_sheet(ws, widths: list[int]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    thin = Side(style="thin", color=BORDER)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.row == 1:
                cell.fill = PatternFill("solid", fgColor=GREEN)
                cell.font = Font(color="FFFFFF", bold=True)
            else:
                cell.font = Font(color="17201D")
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def write_rows(ws, rows: list[list[object]], widths: list[int]) -> None:
    for row in rows:
        ws.append(row)
    style_sheet(ws, widths)


def fill_status(ws, status_col: int) -> None:
    for row in range(2, ws.max_row + 1):
        value = str(ws.cell(row=row, column=status_col).value or "")
        fill = GREEN_LIGHT
        if "待扩展" in value or "注意" in value:
            fill = AMBER_LIGHT
        if "失败" in value or "风险" in value:
            fill = RED_LIGHT
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill)


def main() -> None:
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "交付总览"
    write_rows(
        ws,
        [
            ["项目", "内容", "状态", "备注"],
            ["项目名称", "考试 AI 题库与错题教练", "新增完成", "本地可运行前端应用"],
            ["项目路径", str(BASE_DIR), "新增完成", "全部相关内容已放入该文件夹"],
            ["本地入口", "http://127.0.0.1:5173", "验证通过", "Vite 服务已启动，HTTP 200"],
            ["概念图", str(BASE_DIR / "design" / "exam-ai-question-bank-concept.png"), "新增完成", "用于界面方向核查"],
            ["桌面截图", str(BASE_DIR / "output" / "playwright" / "desktop-dashboard.png"), "验证通过", "1440x1000"],
            ["错题截图", str(BASE_DIR / "output" / "playwright" / "after-wrong-submit.png"), "验证通过", "提交错误答案后生成"],
            ["移动截图", str(BASE_DIR / "output" / "playwright" / "mobile-dashboard.png"), "验证通过", "390x844"],
        ],
        [18, 72, 16, 38],
    )
    fill_status(ws, 3)

    ws = wb.create_sheet("功能清单")
    write_rows(
        ws,
        [
            ["模块", "已实现内容", "交互方式", "状态"],
            ["题库筛选", "按科目、难度、题型、关键词筛选题目", "分段按钮和搜索框", "新增完成"],
            ["答题判分", "支持单选、多选，提交后显示标准答案与解析", "点击选项后提交", "新增完成"],
            ["错题队列", "答错题自动进入优先复盘列表", "点击错题可跳转原题", "新增完成"],
            ["本地 AI 教练", "根据漏选、误选、陷阱和考点生成诊断建议", "提交后刷新诊断", "新增完成"],
            ["掌握度统计", "展示练习量、正确数、待复盘错题、科目进度", "随答题记录变化", "新增完成"],
            ["今日计划", "将当前题或错题加入强化路径", "加入计划并跳转练习", "新增完成"],
            ["真实模型接入", "可扩展到大模型讲解、拍照搜题和自动组卷", "需另行配置接口", "待扩展-需消费确认"],
        ],
        [20, 56, 34, 22],
    )
    fill_status(ws, 4)

    ws = wb.create_sheet("文件清单")
    write_rows(
        ws,
        [
            ["路径", "用途", "状态"],
            ["package.json", "依赖与运行脚本", "新增完成"],
            ["src/data/questions.ts", "样例题库数据与题型定义", "新增完成"],
            ["src/lib/coach.ts", "本地错题教练规则与判分辅助函数", "新增完成"],
            ["src/App.tsx", "主界面、筛选、答题、错题队列和学习计划状态", "新增完成"],
            ["src/styles.css", "桌面与移动端样式", "新增完成"],
            ["README.md", "项目说明，新增内容已用绿色标记", "新增完成"],
            ["docs/交付说明.md", "交付说明，新增内容已用绿色标记", "新增完成"],
            ["output/playwright/*.png", "浏览器验证截图", "验证通过"],
            ["output/reports/*.xlsx", "WPS/Excel 交付说明表", "验证通过"],
        ],
        [36, 58, 20],
    )
    fill_status(ws, 3)

    ws = wb.create_sheet("验证记录")
    write_rows(
        ws,
        [
            ["验证项", "方法", "结果", "证据"],
            ["依赖安装", "npm install", "验证通过", "73 个包审计，0 个漏洞"],
            ["生产构建", "npm run build", "验证通过", "tsc --noEmit 与 vite build 均通过"],
            ["本地服务", "Invoke-WebRequest http://127.0.0.1:5173", "验证通过", "HTTP 200"],
            ["错题流程", "Playwright 点击 A 选项并提交", "验证通过", "显示回答错误，标准答案 B"],
            ["教练诊断", "读取 .coach-summary 文本", "验证通过", "识别漏选 B 与误选 A"],
            ["搜索筛选", "搜索 Python", "验证通过", "筛选结果 1 条"],
            ["移动端", "Playwright 390x844 截图", "验证通过", "无明显遮挡或溢出"],
            ["内置浏览器面板", "Playwright MCP Bridge", "注意", "桥接扩展超时，已使用本地 Playwright 兜底"],
        ],
        [20, 44, 18, 58],
    )
    fill_status(ws, 3)

    ws = wb.create_sheet("专业名词简述")
    write_rows(
        ws,
        [
            ["专业名词", "简述"],
            ["题库", "按考试、章节、题型和难度组织的一组练习题数据。"],
            ["错题本", "记录答错题目、错误选项、标准答案和复盘状态的学习模块。"],
            ["错因诊断", "根据答题差异推断错误来源，例如漏读条件、误选干扰项或公式使用错误。"],
            ["掌握度", "用练习记录估算某科目或知识点当前掌握水平的指标。"],
            ["强化路径", "把薄弱题目按优先级放入复习计划，形成下一步练习队列。"],
            ["Vite", "前端开发与构建工具，负责本地热更新和生产打包。"],
            ["React", "用于构建交互式用户界面的前端框架。"],
            ["Playwright", "浏览器自动化工具，用于截图、点击和验证页面行为。"],
            ["本地规则教练", "不调用云端模型，仅根据代码规则生成学习建议，因此不会产生模型费用。"],
        ],
        [22, 86],
    )
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=BLUE_LIGHT if row % 2 == 0 else GRAY_LIGHT)

    wb.save(REPORT_PATH)

    verified = load_workbook(REPORT_PATH, read_only=True)
    payload = {
        "path": str(REPORT_PATH),
        "sheets": verified.sheetnames,
        "overview_a2": verified["交付总览"]["A2"].value,
        "terms_count": verified["专业名词简述"].max_row - 1,
    }
    verified.close()
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
