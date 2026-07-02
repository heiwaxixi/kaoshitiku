from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]
REPORT_PATH = BASE_DIR / "output" / "reports" / "考试AI题库_双端执行记录_20260702.xlsx"


def style(ws, widths: list[int], status_col: int | None = None) -> None:
    header_fill = PatternFill("solid", fgColor="0F8F69")
    ok_fill = PatternFill("solid", fgColor="DDF7EC")
    note_fill = PatternFill("solid", fgColor="FFF4DE")
    blue_fill = PatternFill("solid", fgColor="EAF2FF")
    thin = Side(style="thin", color="D9E4E0")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
            else:
                cell.font = Font(color="17201D")
                if status_col is None:
                    cell.fill = blue_fill if cell.row % 2 == 0 else ok_fill
                else:
                    status = str(ws.cell(row=cell.row, column=status_col).value or "")
                    cell.fill = note_fill if "说明" in status or "限制" in status else ok_fill
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def append_rows(ws, rows: list[list[object]], widths: list[int], status_col: int | None = None) -> None:
    for row in rows:
        ws.append(row)
    style(ws, widths, status_col)


def main() -> None:
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "双端执行总览"
    append_rows(
        ws,
        [
            ["端", "执行方式", "结果", "证据路径", "说明"],
            ["桌面端", "打开 http://127.0.0.1:5173 并用 Playwright 点击 B 选项后提交", "验证通过", str(BASE_DIR / "output" / "playwright" / "execute-desktop.png"), "显示回答正确，标准答案 B"],
            ["iOS端", "Windows 环境使用 Playwright iPhone 15 Pro 视口执行", "验证通过", str(BASE_DIR / "output" / "playwright" / "execute-ios-iphone.png"), "真实 iOS Simulator 需要 macOS/Xcode；本次为 iPhone Safari 视口验证"],
            ["本地服务", "Get-NetTCPConnection + HTTP 请求", "验证通过", "http://127.0.0.1:5173", "5173 端口正在监听"],
            ["生产构建", "npm run build", "验证通过", str(BASE_DIR / "dist"), "TypeScript 与 Vite 构建通过"],
        ],
        [16, 56, 16, 76, 48],
        3,
    )

    ws = wb.create_sheet("执行结果")
    append_rows(
        ws,
        [
            ["检查项", "桌面端", "iOS端", "结论"],
            ["页面可打开", "已打开本地页面", "iPhone 视口可加载", "验证通过"],
            ["答题交互", "选择 B 并提交", "选择 B 并提交", "验证通过"],
            ["判分结果", "回答正确 / 标准答案 B", "回答正确 / 标准答案 B", "验证通过"],
            ["错题教练", "显示关键规则识别正确", "教练面板可见并显示建议", "验证通过"],
            ["布局", "三栏桌面工作台", "纵向移动端工作流", "验证通过"],
        ],
        [24, 38, 38, 20],
        4,
    )

    ws = wb.create_sheet("专业名词简述")
    append_rows(
        ws,
        [
            ["专业名词", "简述"],
            ["桌面端", "电脑浏览器中的大屏布局，适合三栏并排展示题目、教练和错题队列。"],
            ["iOS端", "面向 iPhone 的移动端布局；当前在 Windows 上通过 iPhone 视口模拟验证。"],
            ["视口", "浏览器可见区域尺寸，用来验证页面在不同屏幕上的排版。"],
            ["Playwright", "浏览器自动化工具，可模拟点击、提交、截图和移动端设备参数。"],
            ["Vite", "前端开发服务器和构建工具，提供本地运行地址和生产打包。"],
            ["响应式布局", "同一套网页根据屏幕宽度自动调整排版，例如桌面三栏、手机单栏。"],
            ["iOS Simulator", "苹果官方 iOS 模拟器，需要 macOS 和 Xcode，Windows 无法原生运行。"],
        ],
        [24, 90],
    )

    wb.save(REPORT_PATH)
    verified = load_workbook(REPORT_PATH, read_only=True)
    payload = {
        "path": str(REPORT_PATH),
        "sheets": verified.sheetnames,
        "a2": verified["双端执行总览"]["A2"].value,
        "terms_count": verified["专业名词简述"].max_row - 1,
    }
    verified.close()
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
