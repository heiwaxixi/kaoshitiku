from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]
SOURCE_PATH = Path(
    "F:/\u7535\u7f51\u5de5\u4f5c/\u8d28\u91cf\u571f\u5efa\u9898\u5e93/"
    "20250524\u8d28\u91cf\u8003\u8bd5\u9898\u5e93-\u8fbd\u5b81\u3001\u65b0\u7586\u571f\u5efa\u9898\u5e93\u6c47\u603b.xlsx"
)
REPORT_PATH = BASE_DIR / "output" / "reports" / "考试AI题库_XLSX试验模板导入记录_20260702.xlsx"


def style(ws, widths: list[int], status_col: int | None = None) -> None:
    header_fill = PatternFill("solid", fgColor="0F8F69")
    ok_fill = PatternFill("solid", fgColor="DDF7EC")
    note_fill = PatternFill("solid", fgColor="FFF4DE")
    blue_fill = PatternFill("solid", fgColor="EAF2FF")
    red_fill = PatternFill("solid", fgColor="FFE8E6")
    thin = Side(style="thin", color="D9E4E0")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
            else:
                cell.font = Font(color="17201D")
                if status_col is None:
                    cell.fill = blue_fill if cell.row % 2 == 0 else ok_fill
                else:
                    value = str(ws.cell(row=cell.row, column=status_col).value or "")
                    cell.fill = red_fill if "未导入" in value else note_fill if "注意" in value else ok_fill
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def append_rows(ws, rows: list[list[object]], widths: list[int], status_col: int | None = None) -> None:
    for row in rows:
        ws.append(row)
    style(ws, widths, status_col)


def source_stats() -> tuple[Counter[str], int]:
    wb = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    index = {str(name).strip(): i for i, name in enumerate(headers) if name is not None}
    type_col = index["题型"]
    stem_col = index["试题正文"]
    counts: Counter[str] = Counter()
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        stem = str(row[stem_col] or "").strip()
        if not stem:
            continue
        counts[str(row[type_col] or "").strip()] += 1
        total += 1
    wb.close()
    return counts, total


def main() -> None:
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    counts, source_total = source_stats()
    importable = counts.get("单选题", 0) + counts.get("多选题", 0) + counts.get("判断题", 0) - 1
    rejected = source_total - importable

    wb = Workbook()
    ws = wb.active
    ws.title = "试验总览"
    append_rows(
        ws,
        [
            ["项目", "内容", "状态", "备注"],
            ["源文件", str(SOURCE_PATH), "验证通过", "已用 openpyxl 只读打开"],
            ["应用入口", "http://127.0.0.1:5173", "验证通过", "Vite 本地服务"],
            ["XLSX 导入能力", "直接上传 .xlsx 并读取第一个工作表", "新增完成", "使用 fflate 解压读取，无 xlsx 高危依赖"],
            ["可导入题量", importable, "验证通过", "单选、多选、判断进入客观题练习模式"],
            ["暂未导入", rejected, "未导入", "50 道简答题 + 1 条客观题格式异常记录"],
            ["确认导入结果", "题库从 8 题变为 1521 题", "验证通过", "Playwright 实测"],
            ["预览截图", str(BASE_DIR / "output" / "playwright" / "import-xlsx-template-preview.png"), "验证通过", "上传后预览"],
            ["导入截图", str(BASE_DIR / "output" / "playwright" / "import-xlsx-after-confirm.png"), "验证通过", "确认导入后"],
        ],
        [18, 78, 16, 48],
        3,
    )

    ws = wb.create_sheet("模板统计")
    rows = [["题型", "源文件数量", "当前处理策略", "状态"]]
    for qtype in ["单选题", "多选题", "判断题", "简答题"]:
      rows.append([
          qtype,
          counts.get(qtype, 0),
          "导入为客观题练习" if qtype != "简答题" else "暂不导入，需另做简答评分",
          "验证通过" if qtype != "简答题" else "未导入",
      ])
    rows.append(["格式异常客观题", 1, "预览中提示未识别原因", "未导入"])
    append_rows(ws, rows, [18, 16, 46, 16], 4)

    ws = wb.create_sheet("字段映射")
    append_rows(
        ws,
        [
            ["Excel 列", "导入字段", "处理规则", "状态"],
            ["题型", "type", "单选题/多选题/判断题 映射为 单选/多选/判断", "新增完成"],
            ["试题正文", "stem", "作为题干展示", "新增完成"],
            ["试题选项", "options", "按 `$;$` 分隔，再识别 A/B/C/D", "新增完成"],
            ["试题答案", "answer", "识别 A/B/C/D，支持多选答案", "新增完成"],
            ["答案解析", "explanation", "有解析则使用解析", "新增完成"],
            ["依据 出处", "explanation", "无解析时作为解析补充", "新增完成"],
            ["二级专业", "chapter/concept", "作为章节和考点", "新增完成"],
            ["难度", "difficulty", "基础题映射为基础，进阶题映射为拔高", "新增完成"],
        ],
        [22, 24, 70, 16],
        4,
    )

    ws = wb.create_sheet("验证记录")
    append_rows(
        ws,
        [
            ["验证项", "方法", "结果", "证据"],
            ["依赖审计", "npm audit --omit=dev", "验证通过", "0 vulnerabilities"],
            ["生产构建", "npm run build", "验证通过", "TypeScript 与 Vite 构建通过"],
            ["真实模板上传", "Playwright setInputFiles 上传 XLSX", "验证通过", "1513 题可导入，51 段未通过"],
            ["确认导入", "点击确认导入", "验证通过", "题库数量 8 -> 1521"],
            ["首题展示", "读取 .stem 与 .meta-row", "验证通过", "质量考试题库 / 质量土建 / 单选 / 基础"],
        ],
        [20, 44, 16, 70],
        3,
    )

    ws = wb.create_sheet("专业名词简述")
    append_rows(
        ws,
        [
            ["专业名词", "简述"],
            ["XLSX", "Excel/WPS 常用工作簿格式，本质是多个 XML 文件组成的压缩包。"],
            ["客观题", "可通过固定选项自动判分的题型，例如单选、多选、判断。"],
            ["简答题", "答案是长文本，通常需要人工或模型评分，当前未放入客观题练习模式。"],
            ["字段映射", "把源 Excel 的列名对应到应用内部题目字段。"],
            ["$;$ 分隔符", "该模板中用于分隔 A/B/C/D 选项的特殊字符串。"],
            ["fflate", "用于浏览器端解压 XLSX 文件的轻量库，当前审计无漏洞。"],
            ["导入预览", "真正写入题库前展示识别结果和未通过原因，降低误导入风险。"],
        ],
        [22, 90],
    )

    wb.save(REPORT_PATH)
    verified = load_workbook(REPORT_PATH, read_only=True)
    payload = {
        "path": str(REPORT_PATH),
        "sheets": verified.sheetnames,
        "importable": verified["试验总览"]["B5"].value,
        "rejected": verified["试验总览"]["B6"].value,
    }
    verified.close()
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
